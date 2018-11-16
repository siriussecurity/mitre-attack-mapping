import os
import sys
import argparse
import simplejson
from openpyxl import load_workbook
from attackcti import attack_client


class MitreAttackMapping():
    """
    MitreAttackMapping is functionality to map your datasources and detections to the MITRE framework and to generate
    layer files that can be loaded into the MITRE ATT&CK Navigator at https://mitre.github.io/attack-navigator/enterprise/

    Start with the mitre-mapping.xlsx file to fill in your organisation's datasources and detections. Then run this
    python script and mitrenize your organisation.
    """

    def __init__(self, mapping_filename):
        """
        Constructor of the MitreMapping class
        :param mapping_filename: filename reference for the Excel file which holds the datasource mapping and query mapping.
        """
        self.mapping_filename = mapping_filename
        self.mitre = attack_client()

        self._get_all_mitre_info()

    def get_mitre_techniques(self):
        """
        :return: Returns all mitre techniques.
        """
        return self.techniques_dict

    def get_mitre_datasources(self, techniques):
        """
        :param techniques: Returns all datasources used in mitre techniques.
        :return:
        """
        return self.datasources

    def generate_layer_files(self):
        """
        Perform the mapping between my datasources and detections from the Excelfile and map it to the MITRE
        datasources and techniques.
        :return:
        """
        self._load_my_datasources_from_file()
        self._load_my_detected_techniques()

        # Do the mapping between my datasources and MITRE datasources:
        for name, my_ds in self.my_datasources.items():
            # First get the techniques that matches the datasources:
            my_techniques = self._colorize_techniques(my_ds, self.detected_techniques[name])

            layer = self._get_layer_template(name)
            layer['techniques'] = my_techniques

            json_string = simplejson.dumps(layer).replace('}, ', '},\n')

            with open(self._normalize_name_to_filename(name) + '.json', 'w') as f:
                f.write(json_string)

    def _get_all_mitre_info(self):
        self.techniques_list = self.mitre.get_all_enterprise_techniques()

        # Convert list to dictionary with technique_id as key:
        self.techniques_dict = {}
        for t in self.techniques_list:
            self.techniques_dict[t['technique_id']] = t

        self.datasources = set()
        for t in self.techniques_dict.values():
            if t['data_sources']:
                for d in t['data_sources']:
                    self.datasources.add(d)

    def _load_my_datasources_from_file(self):
        """
        Getting the datasources per topic from the Excel file.
        :return:
        """
        worksheet_name = 'Datasources'
        wb = load_workbook(os.path.join(self.mapping_filename), read_only=True)
        if worksheet_name not in wb.sheetnames:
            print('No worksheet with name "%s"' % worksheet_name)
            sys.exit(0)

        ws = wb[worksheet_name]
        self.my_datasources = {}
        for col in range(2, ws.max_column+1):
            layer_name = ws.cell(1, col).value
            self.my_datasources[layer_name] = set()
            for row in range(2, ws.max_row+1):
                value = ws.cell(row, col).value
                datasource = ws.cell(row, 1).value
                if value == 'x':
                    self.my_datasources[layer_name].add(datasource)

    def _load_my_detected_techniques(self):
        """
        Getting the techniques being detected per topic from the Excel file.
        :return:
        """
        worksheet_name = 'Detections'
        wb = load_workbook(os.path.join(self.mapping_filename), read_only=True)
        if worksheet_name not in wb.sheetnames:
            print('No worksheet with name "%s"' % worksheet_name)
            sys.exit(0)
        ws = wb[worksheet_name]
        self.detected_techniques = {}
        for col in range(2, ws.max_column + 1):
            layer_name = ws.cell(1, col).value
            self.detected_techniques[layer_name] = set()
            for row in range(2, ws.max_row + 1):
                value = ws.cell(row, col).value
                if value is not None:
                    techniques = value.split(',')
                    for t in techniques:
                        self.detected_techniques[layer_name].add(t)

    def _colorize_techniques(self, my_ds, detected_techniques):
        """
        Determine the color of the techniques based on how many datasources are available per technique and also if
        detection is in place for that technique.
        :param my_ds: The datasources for a specific topic
        :param detected_techniques: The detections for that same specific topic
        :return: A dictionary with techniques that can be used in the layer's output file.
        """
        technique_colors = {}

        # Datasource colors
        c25 = '#f9f1c6'
        c50 = '#ffe766'
        c75 = '#ffd466'
        c99 = '#f6b922'
        c100 = '#c39217'

        # Detection colors
        dc25 = '#bbfcd5'
        dc50 = '#96f2bb'
        dc75 = '#63eb99'
        dc99 = '#33de77'
        dc100 = '#06c452'

        # Color all the techniques based on how many datasources are available and color the techniques based on the
        # detection which are in place.
        for t_id, t in self.techniques_dict.items():
            if t['data_sources']:
                total_ds_count = len(t['data_sources'])
                ds_count = 0
                for ds in t['data_sources']:
                    if ds in my_ds:
                        ds_count += 1
                if total_ds_count > 0:
                    result = (float(ds_count) / float(total_ds_count)) * 100

                    # Check if there is detection in place:
                    if t_id in detected_techniques:
                        color = dc25 if result <= 25 else dc50 if result <= 50 else dc75 if result <= 75 else dc99 if result <= 99 else dc100
                    else:
                        # If not detection is in place, determine the color based on the number of datasources available:
                        color = c25 if result <= 25 else c50 if result <= 50 else c75 if result <= 75 else c99 if result <= 99 else c100

                    technique_colors[t_id] = color
                else:
                    # Appearantly T1205, T1120, T1164 and T1144 don't have a datasources in the API so use a marking color:
                    technique_colors[t_id] = '#dc1a33'

        # Generate a list with techniques to be used in the layer output file.
        my_techniques = []
        my_techniques_set = set()
        for i_ds in my_ds:
            # Loop through all techniques, to find techniques using that datasource:
            for t_id, t in self.techniques_dict.items():
                # If your datasource is in the list of datasources for this technique AND if the
                # technique isn't added yet (by an other datasource):
                if t['data_sources'] and i_ds in t['data_sources'] and t_id not in my_techniques_set:
                    my_techniques_set.add(t_id)

                    for tactic in t['tactic']:
                        d = {}
                        d['techniqueID'] = t_id
                        # d['score'] = 50
                        d['color'] = technique_colors[t_id]
                        d['comment'] = ''
                        d['enabled'] = True
                        d['tactic'] = tactic.lower().replace(' ', '-')
                        my_techniques.append(d)

        return my_techniques

    def _normalize_name_to_filename(self, name):
        """
        Normalize the input filename to a lowercase filename and replace spaces with dashes.
        :param name: input filename
        :return: normalized filename
        """
        return name.lower().replace(' ', '-')

    def _get_layer_template(self, name):
        """
        This returns a basic template that can be loaded into the MITRE ATT&CK navigator.
        :param name: Name (title) of the layer
        :return: A dictionary with a basic layer configuration.
        """
        layer = {}
        layer['name'] = name
        layer['version'] = '2.0'
        layer['domain'] = 'mitre-enterprise'
        layer['description'] = ''
        layer['filters'] = {'stages': ['act'], 'platforms': ['windows', 'linux', 'mac']}
        layer['sorting'] = 0
        layer['viewMode'] = 0
        layer['hideDisable'] = False
        layer['techniques'] = []
        layer['gradient'] = {'colors': ['#ff6666', '#ffe766', '#8ec843'], 'minValue': 0, 'maxValue': 100}
        layer['legendItems'] = []
        layer['showTacticRowBackground'] = False
        layer['tacticRowBackground'] = '#dddddd'
        layer['selectTechniquesAcrossTactics'] = True
        return layer


if __name__ == '__main__':
    argparser = argparse.ArgumentParser()
    argparser.add_argument('-f', dest='MAPPING_FILE', default='mitre-mapping.xlsx', help='Data')
    args = argparser.parse_args()

    m = MitreAttackMapping(args.MAPPING_FILE)

    m.generate_layer_files()
    print('Files written')